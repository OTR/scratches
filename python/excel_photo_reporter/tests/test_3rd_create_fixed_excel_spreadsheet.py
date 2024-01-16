"""The third test case

Create a fixed Excel file from a given broken one according to the task
"""

import os
import re
import unittest

import openpyxl
from openpyxl.utils.cell import column_index_from_string as _from_string

CAN_DRAW = False

try:
    from PIL import Image, ImageDraw, ImageFont
    from io import BytesIO
    CAN_DRAW = True
except ImportError:
    print("We can't draw a picture but it's OK. Install `Pillow` package if you want")

CWD = os.path.dirname(os.path.abspath(__file__))
PATH_TO_BROKEN_SHEET = os.path.join(CWD, "input", "3rd_test_case", "data.xlsx")
PATH_TO_SOURCE_IMAGE = os.path.join(CWD, "input", "3rd_test_case", "parcel.jpg")
PATH_TO_FIXED_SHEET = os.path.join(CWD, "output", "3rd_test_case", "fixed_data.xlsx")
PATH_TO_DEST_IMAGE_FOLDER = os.path.join(CWD, "output", "3rd_test_case", "images")

HEADER = (
    "№",
    "Адрес",
    "Штрих-код",
    "Отредактированный Штрих-код",
    "Фото"
)

BARCODE_PATTERN = re.compile(r"(\d{18})_at_\d{1,2}-\d\.jpg")


class TestCreateFixedFile(unittest.TestCase):
    """"""
    binary_data = bytes()

    @classmethod
    def setUpClass(cls):
        """Load an image in memory"""
        fd = os.open(PATH_TO_SOURCE_IMAGE, os.O_RDONLY | os.O_BINARY)
        with os.fdopen(fd, "rb") as fo:
            cls.binary_data = fo.read()

    def test_01_fix_broken_sheet(self):
        """Задание №1"""
        broken_wb = openpyxl.load_workbook(PATH_TO_BROKEN_SHEET)
        broken_ws = broken_wb.active

        fixed_wb = openpyxl.Workbook()
        fixed_ws = fixed_wb.active

        fixed_ws.append(HEADER)

        for row in broken_ws.iter_rows(min_col=_from_string("A"),
                                       max_col=_from_string("A"),
                                       min_row=broken_ws.min_row,
                                       max_row=broken_ws.max_row):
            broken_barcode = row[0].value
            entry_number = row[0].row

            fixed_row = self._prepare_fixed_row(entry_number, broken_barcode)
            fixed_ws.append(fixed_row)

        # Change style of columns
        fixed_ws.column_dimensions["C"].width = 25.0
        fixed_ws.column_dimensions["D"].width = 30.0

        # Save on a disk
        try:
            fixed_wb.save(PATH_TO_FIXED_SHEET)
            print(f"Task №1 is done. Take a look at {PATH_TO_FIXED_SHEET} file")
        except PermissionError as err:
            if err.args[0] == 13:
                self.fail("File is already opened in another application. Close it!")

    @staticmethod
    def _prepare_fixed_row(entry_number, broken_barcode):
        """Helper function"""
        fixed_barcode = broken_barcode.lstrip("0")

        # Get address by entry number
        quotient, remainder = divmod(entry_number - 1, 69)
        if quotient % 2 == 0:  # Even
            cell_num = remainder + 1
        else:  # Odd
            cell_num = 69 - remainder
        tier_num = quotient + 1
        address = f"{cell_num}-{tier_num}"

        return tuple([entry_number, address, broken_barcode, fixed_barcode])

    def test_21_create_image(self):
        """Задание №2.1"""
        wb = openpyxl.load_workbook(PATH_TO_FIXED_SHEET)
        ws = wb.active

        for row in ws.iter_rows(min_col=_from_string("B"),
                                max_col=_from_string("D"),
                                min_row=ws.min_row + 1,  # Skip header
                                max_row=ws.max_row):

            # Create images
            address = row[0].value  # B
            fixed_barcode = row[-1].value  # D
            if fixed_barcode.isdecimal() and len(fixed_barcode) == 18:
                path_to_image_with_barcode = os.path.join(PATH_TO_DEST_IMAGE_FOLDER,
                                                          f"{fixed_barcode}_at_{address}.jpg")
                initial_image = type(self).binary_data
                # if we have Pillow library then
                if CAN_DRAW:
                    # Draw a barcode
                    stream = BytesIO(initial_image)
                    image = Image.open(stream)
                    draw = ImageDraw.Draw(image)
                    font = ImageFont.truetype("arial.ttf", 64)
                    draw.text((150, 900), fixed_barcode, fill="black", font=font)
                    new_stream = BytesIO()
                    image.save(new_stream, format="JPEG")
                    initial_image = new_stream.getvalue()

                fd = os.open(path_to_image_with_barcode, os.O_WRONLY | os.O_BINARY | os.O_CREAT)
                with os.fdopen(fd, "wb") as fo:
                    fo.write(initial_image)

        print(f"Task №2.1 is done. Take a look at {PATH_TO_DEST_IMAGE_FOLDER} folder")

    def test_22_update_sheet(self):
        """Задание №2.2"""
        # Find all images with barcode in their name in given folder
        goods_found = {}
        for root, dirs, files in os.walk(PATH_TO_DEST_IMAGE_FOLDER):
            for _file in files:
                result = BARCODE_PATTERN.match(_file)
                if result is None:
                    continue
                # if we found images with barcode in given folder
                if result.groups():
                    barcode = result.groups()[0]
                    goods_found[barcode] = os.path.join(root, _file)

        wb = openpyxl.load_workbook(PATH_TO_FIXED_SHEET)
        ws = wb.active

        for row in ws.iter_rows(min_col=_from_string("B"),
                                max_col=_from_string("E"),
                                min_row=ws.min_row + 1,  # Skip header
                                max_row=ws.max_row):

            # Create images
            address = row[0].value  # B
            fixed_barcode = row[-2].value  # D
            cell_with_hyperlink = row[-1]  # E

            if fixed_barcode.isdecimal() and len(fixed_barcode) == 18:
                if fixed_barcode in goods_found.keys():
                    path_to_img = goods_found[fixed_barcode]
                    del goods_found[fixed_barcode]
                    # Create hyperlink
                    cell_with_hyperlink.hyperlink = path_to_img
                    cell_with_hyperlink.value = address
                    cell_with_hyperlink.style = "Hyperlink"

        # Save on a disk
        try:
            wb.save(PATH_TO_FIXED_SHEET)
            print(f"Task №2.2 is done. Take a look at {PATH_TO_FIXED_SHEET} file")
        except PermissionError as err:
            if err.args[0] == 13:
                self.fail("File is already opened in another application. Close it!")


if __name__ == "__main__":
    unittest.main()
