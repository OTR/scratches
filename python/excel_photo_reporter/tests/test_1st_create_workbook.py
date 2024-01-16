"""
The first test case. Check creation of a proper xlsx file.

When call `ws[row][column]`
IndexError: 9 is not a valid coordinate or range [0][0]
IndexError: tuple index out of range [1][1]
ws[1][0] => A1  # Не работает если к ячейке ни разу не обращались

When call `ws.cell(row=1, column=0)`
ValueError: Row or column values must be at least 1
ws.cell(row=1, column=1) => A1

tuple(ws.rows)[0][0] => A1

Вид > Страничный режим - покажет ячейки к которым обращались
"""

import datetime
import pathlib
import unittest
import zipfile
from string import ascii_uppercase

import openpyxl
from openpyxl.utils.cell import column_index_from_string as _fromstring
from openpyxl.utils.cell import coordinate_to_tuple as _totuple

# TODO: move it into __init__.py
CWD = pathlib.Path(__file__).resolve().parent


def closure():
    """Return a rapper over print with enumerated output. ipython-like
    style."""
    counter = 1

    def inner(*args):
        nonlocal counter
        print("Output {counter}".format(counter=counter))
        print(*args)
        print("_" * 80)
        counter += 1

    return inner


myprint = closure()


# TODO: rename according to its behavior
class TestCreateBasicWorkbook(unittest.TestCase):
    """A test case that creates and fills up a spreadsheet."""

    @classmethod
    def setUpClass(cls):
        """FIXME: refactor cls.cwd."""
        cls.cwd = pathlib.Path(__file__).resolve().parent

    def setUp(self):
        """Find out absolute path to the working directory.
        hint: this function is called before each test method.
        setUpClass is called once before all test methods"""
        pass

    def test_creation(self):
        """Make an object of class Workbook."""
        wb = openpyxl.Workbook()  # FIXME: self.wb?
        ws1 = wb.active  # Get active worksheet

        # Fill up some cells
        ws1["A1"], ws1["B1"], ws1["C1"] = 1, 2, 3
        ws1.append([4, 5, 6])  # Append cells starting from next line?
        ws1.append([7, 8, 9])  # Append cells starting from next line?

        ws1.title = "Matrix"

        # All other workbook / worksheet attributes are NOT copied - e.g.
        # Images, Charts.
        if "Matrix" in wb.sheetnames:
            _ws1 = wb.copy_worksheet(wb["Matrix"])
            _ws1.sheet_properties.tabColor = "1072BA"
            _ws1.title = "Matrix hidden copy"

            cells = _ws1["A1":"C3"]
            self.assertEqual(
                sum(cell.value for column in cells for cell in column),
                sum(range(1, 9 + 1)),  # Sum from 1 to 9
            )

        # Create and fill up second sheet
        ws2 = wb.create_sheet("Main Report", 0)
        ws2["A2"] = datetime.datetime.now()
        ws2.column_dimensions["A"].width = 20.0

        # Save on a disk
        # FIXME: use module defined variable
        dest_path = self.cwd / "output" / \
                    "1st_test_case" / "1st_test_create_workbook.xlsx"
        try:
            wb.save(dest_path)
        except PermissionError as err:
            # https://stackoverflow.com/questions/41910583/errno-13-permission-denied-python
            if err.args[0] == 13:
                self.fail("File is already opened in other application.\
                           Close it!")
                # No need to raise Exception cause `fail` method
                # raises failureException which is AssertionError

    def test_loading(self):
        """Load xlsx created with Excel to make sure everything is parsed
        correctly."""
        pass

    def tearDown(self):
        """"""
        pass


class TestForFindingBetterPractices(unittest.TestCase):
    """Just playing around with openpyxl API to find out its pitfalls."""

    @classmethod
    def setUpClass(cls) -> None:
        """Prepare common workbook for all tests. FIXME: create separate
        workbook for each test."""
        cls.wb = openpyxl.Workbook()
        cls.ws = cls.wb.active
        cls.win_path = CWD / "output" / "1st_test_case" / "example_00.xlsx"
        cls.path = cls.win_path.as_posix()

    def test_00_example(self):
        """"""
        self.ws.title = "Репорт"
        myprint(self.wb["Репорт"][1][0].value)  # Output 1

        self.wb["Репорт"][1][0].value = "Hello"  # A1
        self.wb["Репорт"][2][0].value = "World"  # A2
        self.wb["Репорт"][3][0].value = "!"  # A3
        self.wb["Репорт"]["A4"].value = "print"

        self.wb.create_sheet("Отчет", -1)  # In the beginning
        self.wb["Отчет"]["A1"].value = 555
        self.ws["B1"].value = True  # active все ещё страница Репорт
        # Before -1, if name already exist it will name new sheet as Отчет1
        self.wb.create_sheet("Отчет", 0)
        myprint(tuple(self.wb)[0].title)  # Output 2
        for name in self.wb.sheetnames:
            myprint(name)  # Output 3, 4, 5

        for sheet in self.wb:
            myprint(sheet["A1"].value)  # Output 6, 7, 8

        self.ws["A1"] = "foo"
        myprint(self.ws["A1"].value)  # Output 9

        self.ws["A1"].value = "baz"
        myprint(self.ws["A1"].value)  # Output 10

        a1 = self.ws.cell(row=1, column=1, value="bar")
        myprint(self.ws["A1"].value)  # Output 11

        tuple(self.ws.rows)[0][0].value = "bin"
        myprint(self.ws["A1"].value)  # Output 12

    def test_01_example(self):
        """
        Make a table 3 by 3

        1 2 3
        4 5 6
        7 8 9
        """
        myiter = iter(range(1, 10))
        for row in range(1, 3 + 1):  # Rows start from 1
            for column in range(1, 3 + 1):  # Columns start from 1
                myprint(row, column)  # Output 1, ..., 9
                self.ws.cell(row=row, column=column).value = next(myiter)
        myprint(self.ws.min_row, self.ws.min_column)  # 1 1 Output 10
        myprint(self.ws.max_row, self.ws.max_column)  # 3 3 Output 11

        lst = [cell.value for cell in self.ws[1:1]]
        myprint(len(lst))  # 3 Output 12

        self.ws["A1:AA1"]  # Take a slice bigger that cells exist
        lst2 = [cell.coordinate for cell in self.ws[1:1]]
        myprint(len(lst2))  # 27  Output 13
        myprint(max(lst2))  # "Z1" Output 14

        self.ws["A1:AB1"]  # Take a slice bigger that cells exist
        lst3 = [cell.coordinate for cell in self.ws[1:1]]
        myprint(len(lst3))  # 52 Output 15 FIXME: Got 28?
        myprint(max(lst3))  # Still "Z1" Output 16

        lst4 = [_totuple(cell) for cell in lst3]
        max_cell = max(lst4, key=lambda row_col: row_col[1])
        myprint(
            self.ws.cell(row=max_cell[0], column=max_cell[1]).coordinate,
        )  # AB1 Output 17

        myprint(
            max(self.ws[1:1], key=lambda cell: cell.column).coordinate,
        )  # AB1 Output 18

        key = lambda cell: cell.column if cell.value is not None else -1
        myprint(max(self.ws[1:1], key=key).coordinate)  # C1 Output 19

        max_cell2 = max(lst4, key=lambda row_col: row_col[1])

        lst9 = [cell.coordinate for cell in self.ws[1:1] if
                cell.value is not None]
        myprint(len(lst9))  # 3 Output 20
        myprint(max(lst9))  # C1 Output 21

    def test_02_example(self):
        """"""
        myprint(self.ws.max_column)  # 3 Output 1 FIXME: Got 1
        # After accessing non-existing cells max column increases from 3 to 27
        # But still filled with None values
        cell_range = self.ws["A1:AA1"]
        self.ws["AA1"].value = "Gimme"
        self.ws["AA2"].value = "Dat"
        myprint(len(cell_range[0]))  # 27 Output 2
        myprint(len(ascii_uppercase) + 1)  # 27 Output 3
        myprint(_fromstring("AA"))  # 27 Output 4
        # max_column is lying, don't use it to get max not empty cell
        myprint(self.ws.max_column)  # 27 Output 5

        myprint(self.ws["AA3"].value)  # => None Output 6
        # => None Output 7
        myprint(self.ws.cell(row=3, column=_fromstring("AA")).value)

        myprint(self.ws["AA3"].value)  # => None Output 8 AA1
        # => None Output 9 AA1
        myprint(self.ws.cell(row=3, column=_fromstring("AA")).value)
        # Gimme AA1 Output 10
        myprint(self.ws[1][_fromstring("AA") - 1].value)
        # Index error, if call before cell_range line. Bad example
        # AA2 Dat Output 11
        myprint(tuple(self.ws.rows)[1][_fromstring("AA") - 1].value)

        cell_range = self.ws["A1:AA1"]
        another_range = self.ws["A1":"AA1"]
        myprint(cell_range == another_range)  # True Output 12

    def test_zipnames(self):
        """Test `zipnames` method."""
        self.zipnames()

    def zipnames(self):
        """Extract xlsx file (which is a zip archive to given folder."""
        with zipfile.ZipFile(self.win_path, "r") as fd:
            myprint(fd.namelist())
            fd.extractall(self.win_path.parent / "temp")

    def save(self):
        """Save on a disk."""
        try:
            self.wb.save(self.win_path)
        except PermissionError as err:
            if err.errno == 13:
                myprint("This file is already opened in another application. \
                        Close it!")

    def tearDown(self) -> None:
        """Save result on a disk after each test case."""
        # FIXME: save each test in its own file
        self.save()


if __name__ == "__main__":
    unittest.main()
