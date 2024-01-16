"""
The second test case

Create an Excel report about existing books on Amazon by given keyword
(by default keyword is `python`)
"""

import csv
import pathlib
import unittest

import openpyxl
from openpyxl.drawing.image import Image

CWD = pathlib.Path(__file__).resolve().parent


class TestCreateAmazonBooksReport(unittest.TestCase):
    """Get books from a CSV file."""

    def setUp(self):
        """"""
        # TODO: Make sure a CSV file exists
        pass

    def test_create_sheet_with_amazon_books(self):
        """"""
        book_rows = []

        # 3. Load a CSV file in memory
        path_to_csv_file = CWD / "input" / "2nd_test_case" \
                           / "handmade_amazon_books_fields.csv"
        with open(path_to_csv_file, newline='', encoding='utf-8') as csv_file:
            reader = csv.reader(csv_file, delimiter=';')
            for row in reader:
                book_rows.append(row)

        # 4. Create a new WorkBook in memory
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Amazon Book's Report"

        # Fill it up with a dataset from a CSV file
        for row in range(1, len(book_rows) + 1):
            for column in range(1, len(book_rows[0]) + 1):
                _ = ws.cell(column=column,
                            row=row,
                            value=book_rows[row - 1][column - 1])

        # 5. Widen each row up to 165 pt (in Excel measures) except the 1st row
        # (row with titles) which is 57 pt
        ws.row_dimensions[1].height = 57.0
        for row_number in range(2, ws.max_row + 1):  # [2, ..., 16+1]
            ws.row_dimensions[row_number].height = 165.0

        # 6. Widen columns according to the following
        # Position: 5, Title: 36, Author: 11, Price: 11, Date: 11, Cover: 21
        ws.column_dimensions["A"].width = 5.0  # Position
        ws.column_dimensions["B"].width = 36.0  # Title
        ws.column_dimensions["C"].width = 11.0  # Author
        ws.column_dimensions["D"].width = 11.0  # Price
        ws.column_dimensions["E"].width = 11.0  # Date
        ws.column_dimensions["F"].width = 27.0  # Cover

        # 7. Change the 1st row:
        #     * Background color to light blue RGB(155, 194, 230)
        #     * Horizontally align text to center
        #     * Vertically align text to top
        #     * Make a font style bold
        gen = (f"F{x}" for x in range(2, 29))  # TODO: get rid of magic numbers

        img_dir = CWD / "output" / "scrape_books"
        for img_path in img_dir.glob("*.jpg"):
            img = Image(img_path.as_posix())
            ws.add_image(img, next(gen))

        # Save it on a disk
        dest_path = CWD / "output" / "2nd_test_case" / \
                    "2nd_test_create_amazon_book_report.xlsx"
        try:
            wb.save(dest_path)
        except PermissionError as err:
            if err.args[0] == 13:
                self.fail("File is already opened in other application. Close it!")

    def tearDown(self):
        """"""
        # TODO: delete all pictures from scrape_books dir
        pass


if __name__ == "__main__":
    unittest.main()
